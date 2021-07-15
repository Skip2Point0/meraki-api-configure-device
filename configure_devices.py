import requests
import json
from openpyxl import load_workbook

# ################################################# PARAMETERS BELOW ###################################################
# Please fill in parameters below.
# You can add devices to one or multiple networks.
# ap_abr is a variable that is used to track device names. Typical syntax for NN1-AP1-25 is NN1-
# Make sure that every parameter is added in correct order: Network Name1 > Sheet Name1 > NN1-
# ######################################################################################################################
meraki_api = 'APIKEY'
organization_id = 'Organization Name'
network_ids = ['Network Name1', 'Network Name2']
spread = 'spreadsheet.xlsx'
address = ['Physical Address of First Network', 'Physical Address of Second Network']
tabs = ['Sheet Name1', 'Sheet Name2']
ap_abr = ['NN1-', 'NN2-']
# ##################################################### PARAMETERS #####################################################
# If using default spreadsheet, no changes below are necessary.
# Custom spreadsheets require column changes below.
# ######################################################################################################################
ap_name_column = 'A'
notes_column = 'B'
tags_column = 'C'
serials_column = 'D'
mac_column = 'E'
# ################################################# PARAMETERS ABOVE ###################################################
net_dictionary = {}
headers = {
    'X-Cisco-Meraki-API-Key': meraki_api,
    'Content-Type': 'application/json'
}
print(spread)
wb = load_workbook(spread)


def pull_organization_id(head):
    url = "https://api.meraki.com/api/v0/organizations"
    payload = {}

    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    response = json.loads(response)
    for dicti in response:
        name = dicti["name"]
        if name == organization_id:
            org_id = dicti["id"]
            print("#################################################")
            print(name + "\n" + "Organization ID: " + org_id)
            print("#################################################")
            return org_id
        else:
            continue


def pull_organization_networks(head):
    global net_dictionary
    global organization_id

    organization_id = pull_organization_id(head)
    url = "https://api.meraki.com/api/v0/organizations/" + organization_id + "/networks"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    json_response = json.loads(response)
    for networks in json_response:
        name = networks['name']
        n_id = networks['id']
        net_dictionary[name] = n_id
    print(net_dictionary)
    return net_dictionary


def pull_destination_networks():
    global network_ids

    dest_network_ids = []
    for n in network_ids:
        for i in net_dictionary:
            if n == i:
                print("Destination Network: " + n)
                dest_network_ids.append(net_dictionary[n])
                break
            else:
                continue
    print(dest_network_ids)
    return dest_network_ids


pull_organization_networks(headers)
networks_dest = pull_destination_networks()


def meraki_ap_config(workbook, networks, addr, tbs, abr, name, notes, tags, serial, head):
    ap_index = []
    print("########################################")
    print("APPLYING AP CONFIG...")
    incr = 0
    for tab in tbs:
        sheet = workbook[tab]
        row_index = []
        print("########################################")
        print("NETWORK:  " + tab)
        print("########################################")

        incr2 = 1
        for row in range(sheet.max_row):
            a = sheet[name + str(incr2)]
            unsorted_ap = a.value
            incr2 = incr2 + 1

            if abr[incr] in str(unsorted_ap):
                n_ap_index = row + 1
                ap_index.append(unsorted_ap)
                row_index.append(row)
                serials = sheet[serial + str(n_ap_index)].value
                ad = addr[incr]
                tg = sheet[tags + str(n_ap_index)].value
                nt = sheet[notes + str(n_ap_index)].value
                url = "https://api.meraki.com/api/v0/networks/" + networks[incr] + "/devices/" + serials
                print(url)
                payload = {
                    "address": ad,
                    "tags": tg,
                    "name": unsorted_ap,
                    "notes": nt,
                }
                payload = json.dumps(payload)
                print(payload)
                response = requests.request("PUT", url, headers=head, data=payload)
                f = response.content
                print(f)

        incr = incr + 1


meraki_ap_config(wb, networks_dest, address, tabs, ap_abr, ap_name_column, notes_column, tags_column, serials_column,
                 headers)
